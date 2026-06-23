import csv
import io
import json
import os
import tempfile
from decimal import Decimal
import jdatetime
from django.core.management import call_command
from django.db import transaction
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .backup_fixture import build_dumpdata_backup_json, repair_backup_fixture_shamsi_dates
from .models import (
    Student,
    Teacher,
    TeacherSalaryPayment,
    SchoolClass,
    FeeType,
    Payment,
    ExpenseCategory,
    Expense,
)
from .serializers import (
    StudentSerializer,
    TeacherSerializer,
    TeacherSalaryPaymentSerializer,
    SchoolClassSerializer,
    FeeTypeSerializer,
    PaymentSerializer,
    ExpenseCategorySerializer,
    ExpenseSerializer,
)


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all().order_by("-id")
    serializer_class = StudentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.query_params.get("q")
        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(registration_number__icontains=q)
                | Q(father_name__icontains=q)
                | Q(grandfather_name__icontains=q)
                | Q(phone__icontains=q)
            )

        for field in ["name", "registration_number", "father_name", "grandfather_name", "phone"]:
            value = self.request.query_params.get(field)
            if value:
                lookup = {f"{field}__icontains": value}
                qs = qs.filter(**lookup)
        return qs

    @action(detail=True, methods=["get"])
    def payments(self, request, pk=None):
        student = self.get_object()
        payments = student.payments.all().order_by("-created_at")
        serializer = PaymentSerializer(payments, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def report(self, request, pk=None):
        student = self.get_object()
        payments = student.payments.all()
        totals = (
            payments.values("fee_type__name")
            .annotate(total=Sum("amount"))
            .order_by("fee_type__name")
        )
        total_paid = payments.aggregate(total=Sum("amount")).get("total") or Decimal("0")
        return Response(
            {
                "student_id": student.id,
                "total_paid": total_paid,
                "totals_by_fee_type": list(totals),
            }
        )

    @action(detail=False, methods=["post"], url_path="import")
    def import_students(self, request):
        upload = request.FILES.get("file")
        mode = (request.data.get("mode") or "partial").strip().lower()
        if not upload:
            return Response({"detail": "File is required."}, status=status.HTTP_400_BAD_REQUEST)
        if mode not in {"partial", "strict"}:
            return Response({"detail": "mode must be 'partial' or 'strict'."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = _read_student_rows(upload)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({"detail": "No data rows found in file."}, status=status.HTTP_400_BAD_REQUEST)

        classes_by_id = {str(c.id): c for c in SchoolClass.objects.all()}
        classes_by_name_year = {
            (c.name.strip().lower(), c.year_shamsi.strip()): c for c in SchoolClass.objects.all()
        }

        errors = []
        students_to_create = []
        existing_registration_numbers = {
            value
            for value in Student.objects.exclude(registration_number="").values_list("registration_number", flat=True)
        }
        pending_registration_numbers = set()
        max_rows = 10000
        if len(rows) > max_rows:
            return Response(
                {"detail": f"Too many rows ({len(rows)}). Max allowed is {max_rows}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for row_number, row in rows:
            student_obj, row_errors = _build_student_from_row(row, row_number, classes_by_id, classes_by_name_year)
            if row_errors:
                errors.extend(row_errors)
                continue
            reg_no = (student_obj.registration_number or "").strip()
            if reg_no in existing_registration_numbers or reg_no in pending_registration_numbers:
                errors.append(
                    {
                        "row": row_number,
                        "field": "registration_number",
                        "message": "Registration number already exists.",
                    }
                )
                continue
            pending_registration_numbers.add(reg_no)
            students_to_create.append(student_obj)

        if mode == "strict" and errors:
            return Response(
                {
                    "total_rows": len(rows),
                    "imported": 0,
                    "failed": len(errors),
                    "errors": errors[:200],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            Student.objects.bulk_create(students_to_create, batch_size=500)

        return Response(
            {
                "total_rows": len(rows),
                "imported": len(students_to_create),
                "failed": len(errors),
                "errors": errors[:200],
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        headers = [
            "class_id",
            "class_name",
            "year_shamsi",
            "name",
            "registration_number",
            "father_name",
            "grandfather_name",
            "phone",
            "monthly_fee_override",
            "transport_fee_override",
            "uniform_fee_override",
            "book_fee_override",
            "previous_balance",
        ]
        sample = "1,,1404,Ali,REG-001,Reza,Hassan,700000001,1200,0,0,0,3500"
        content = ",".join(headers) + "\n" + sample + "\n"
        response = HttpResponse(content, content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="students_import_template.csv"'
        return response


class TeacherViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.all().order_by("-id")
    serializer_class = TeacherSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.query_params.get("q")
        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(father_name__icontains=q)
                | Q(phone__icontains=q)
                | Q(email__icontains=q)
                | Q(department__icontains=q)
            )
        return qs


class TeacherSalaryPaymentViewSet(viewsets.ModelViewSet):
    queryset = TeacherSalaryPayment.objects.select_related("teacher").all().order_by("-date_shamsi", "-id")
    serializer_class = TeacherSalaryPaymentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        teacher_id = self.request.query_params.get("teacher_id")
        month = self.request.query_params.get("month")
        start = self.request.query_params.get("start")
        end = self.request.query_params.get("end")

        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        if month:
            qs = qs.filter(month_shamsi=month)
        if start and end:
            try:
                start_date = _parse_shamsi_date(start)
                end_date = _parse_shamsi_date(end)
            except ValueError:
                return qs.none()
            qs = qs.filter(date_shamsi__gte=start_date, date_shamsi__lte=end_date)
        return qs


class SchoolClassViewSet(viewsets.ModelViewSet):
    queryset = SchoolClass.objects.all().order_by("-year_shamsi", "name")
    serializer_class = SchoolClassSerializer


class FeeTypeViewSet(viewsets.ModelViewSet):
    queryset = FeeType.objects.all().order_by("name")
    serializer_class = FeeTypeSerializer


class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.select_related("student", "fee_type").all().order_by("-created_at")
    serializer_class = PaymentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get("student_id")
        fee_type_id = self.request.query_params.get("fee_type_id")
        start = self.request.query_params.get("start")
        end = self.request.query_params.get("end")
        month = self.request.query_params.get("month")

        if student_id:
            qs = qs.filter(student_id=student_id)
        if fee_type_id:
            qs = qs.filter(fee_type_id=fee_type_id)
        if month:
            qs = qs.filter(month_shamsi=month)
        if start and end:
            try:
                start_date = _parse_shamsi_date(start)
                end_date = _parse_shamsi_date(end)
            except ValueError:
                return qs.none()
            qs = qs.filter(date_shamsi__gte=start_date, date_shamsi__lte=end_date)
        return qs


class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    queryset = ExpenseCategory.objects.all().order_by("name")
    serializer_class = ExpenseCategorySerializer


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.select_related("category").all().order_by("-created_at")
    serializer_class = ExpenseSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        category_id = self.request.query_params.get("category_id")
        start = self.request.query_params.get("start")
        end = self.request.query_params.get("end")
        month = self.request.query_params.get("month")

        if category_id:
            qs = qs.filter(category_id=category_id)
        if month:
            try:
                year, month_num = _parse_shamsi_month(month)
            except ValueError:
                return qs.none()
            qs = qs.filter(date_shamsi__year=year, date_shamsi__month=month_num)
        if start and end:
            try:
                start_date = _parse_shamsi_date(start)
                end_date = _parse_shamsi_date(end)
            except ValueError:
                return qs.none()
            qs = qs.filter(date_shamsi__gte=start_date, date_shamsi__lte=end_date)
        return qs


class ReportSummaryView(APIView):
    def get(self, request):
        period = request.query_params.get("period")  # day, month, year, custom
        date = request.query_params.get("date")  # YYYY-MM-DD or YYYY-MM or YYYY
        start = request.query_params.get("start")
        end = request.query_params.get("end")
        include_items = request.query_params.get("include_items") == "1"

        payment_qs = Payment.objects.select_related("student", "fee_type", "school_class").all()
        expense_qs = Expense.objects.select_related("category").all()

        if period in {"day", "month", "year"} and date:
            try:
                if period == "day":
                    target_date = _parse_shamsi_date(date)
                    payment_qs = payment_qs.filter(date_shamsi=target_date)
                    expense_qs = expense_qs.filter(date_shamsi=target_date)
                elif period == "month":
                    year, month_num = _parse_shamsi_month(date)
                    payment_qs = payment_qs.filter(date_shamsi__year=year, date_shamsi__month=month_num)
                    expense_qs = expense_qs.filter(date_shamsi__year=year, date_shamsi__month=month_num)
                elif period == "year":
                    year = _parse_shamsi_year(date)
                    payment_qs = payment_qs.filter(date_shamsi__year=year)
                    expense_qs = expense_qs.filter(date_shamsi__year=year)
            except ValueError as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        elif period == "custom" and start and end:
            try:
                start_date = _parse_shamsi_date(start)
                end_date = _parse_shamsi_date(end)
            except ValueError as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
            payment_qs = payment_qs.filter(date_shamsi__gte=start_date, date_shamsi__lte=end_date)
            expense_qs = expense_qs.filter(date_shamsi__gte=start_date, date_shamsi__lte=end_date)

        total_revenue = payment_qs.aggregate(total=Sum("amount")).get("total") or Decimal("0")
        total_expenses = expense_qs.aggregate(total=Sum("amount")).get("total") or Decimal("0")
        profit = total_revenue - total_expenses

        response = {
            "total_revenue": total_revenue,
            "total_expenses": total_expenses,
            "profit": profit,
        }

        if include_items:
            response["payments"] = PaymentSerializer(payment_qs.order_by("-created_at"), many=True).data
            response["expenses"] = ExpenseSerializer(expense_qs.order_by("-created_at"), many=True).data

        return Response(response)


class MonthlyDueFeesView(APIView):
    """
    Returns students who still owe monthly and/or transport fees through a given Shamsi month.

    For each fee type, dues are cumulative from month 01 (Hamal) of that Shamsi year through
    the requested month.

    The breakdown is:
    - *_fee_previous: sum of shortfalls for all months *before* the requested month
    - *_fee_current: shortfall for the requested month itself
    - *_fee: previous + current (backwards‑compatible total)
    - *_previous_months_count: how many of those previous months still have a balance (>0),
      for the bill "برج" column on باقیات rows (number of months, not the month name).
    """

    def get(self, request):
        month_shamsi = request.query_params.get("month_shamsi")
        if not month_shamsi:
            year = request.query_params.get("year")
            month = request.query_params.get("month")
            if year and month:
                try:
                    month_shamsi = f"{int(year):04d}-{int(month):02d}"
                except ValueError as exc:
                    return Response(
                        {"detail": "Invalid year/month format."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        if not month_shamsi:
            return Response(
                {"detail": "month_shamsi is required (YYYY-MM)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            year_int, end_month_int = _parse_shamsi_month(month_shamsi)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        months_in_range = [f"{year_int}-{m:02d}" for m in range(1, end_month_int + 1)]
        target_month = month_shamsi
        dues_from_month_shamsi = f"{year_int:04d}-01"

        class_id = request.query_params.get("class_id")

        monthly_fee_types = FeeType.objects.filter(name__icontains="monthly")
        if not monthly_fee_types.exists():
            return Response(
                {"detail": "Monthly FeeType not found. Create a FeeType with 'monthly' in its name."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        transport_fee_types = FeeType.objects.filter(name__icontains="transport")
        previous_balance_fee_types = _previous_balance_fee_types()

        students_qs = Student.objects.select_related("school_class").filter(school_class__isnull=False)
        if class_id:
            students_qs = students_qs.filter(school_class_id=class_id)

        student_ids = list(students_qs.values_list("id", flat=True))

        monthly_paid_by_student_month = {}
        if student_ids:
            for row in (
                Payment.objects.filter(
                    student_id__in=student_ids,
                    month_shamsi__in=months_in_range,
                    fee_type__in=monthly_fee_types,
                )
                .values("student_id", "month_shamsi")
                .annotate(paid=Sum("amount"))
            ):
                monthly_paid_by_student_month[(row["student_id"], row["month_shamsi"])] = row["paid"]

        transport_paid_by_student_month = {}
        if student_ids:
            for row in (
                Payment.objects.filter(
                    student_id__in=student_ids,
                    month_shamsi__in=months_in_range,
                    fee_type__in=transport_fee_types,
                )
                .values("student_id", "month_shamsi")
                .annotate(paid=Sum("amount"))
            ):
                transport_paid_by_student_month[(row["student_id"], row["month_shamsi"])] = row["paid"]

        previous_balance_paid_by_student = {}
        if student_ids and previous_balance_fee_types.exists():
            for row in (
                Payment.objects.filter(
                    student_id__in=student_ids,
                    month_shamsi__in=months_in_range,
                    fee_type__in=previous_balance_fee_types,
                )
                .values("student_id")
                .annotate(paid=Sum("amount"))
            ):
                previous_balance_paid_by_student[row["student_id"]] = row["paid"] or Decimal("0")

        results = []
        for student in students_qs:
            start_month_shamsi = max(_student_enrolled_month_shamsi(student), dues_from_month_shamsi)
            if start_month_shamsi > target_month:
                student_months_in_range = []
            else:
                student_months_in_range = _iter_shamsi_months(start_month_shamsi, target_month)

            expected_monthly = (
                student.monthly_fee_override
                if student.monthly_fee_override is not None
                else student.school_class.monthly_fee
            )
            expected_transport = (
                student.transport_fee_override
                if student.transport_fee_override is not None
                else student.school_class.transport_fee
            )

            # Totals across the whole Hamal→target period
            paid_monthly_total = Decimal("0")
            paid_transport_total = Decimal("0")

            # Split dues into "previous months" vs "current month"
            due_monthly_previous = Decimal("0")
            due_monthly_current = Decimal("0")
            due_transport_previous = Decimal("0")
            due_transport_current = Decimal("0")
            monthly_previous_unpaid_months = 0
            transport_previous_unpaid_months = 0

            sid = student.id
            for m in student_months_in_range:
                paid_m = monthly_paid_by_student_month.get((sid, m)) or Decimal("0")
                paid_monthly_total += paid_m
                short_m = max(expected_monthly - paid_m, Decimal("0"))
                if m == target_month:
                    due_monthly_current += short_m
                else:
                    due_monthly_previous += short_m
                    if short_m > 0:
                        monthly_previous_unpaid_months += 1

                paid_t = transport_paid_by_student_month.get((sid, m)) or Decimal("0")
                paid_transport_total += paid_t
                short_t = max(expected_transport - paid_t, Decimal("0"))
                if m == target_month:
                    due_transport_current += short_t
                else:
                    due_transport_previous += short_t
                    if short_t > 0:
                        transport_previous_unpaid_months += 1

            due_monthly_total = due_monthly_previous + due_monthly_current
            due_transport_total = due_transport_previous + due_transport_current
            previous_balance_paid = previous_balance_paid_by_student.get(student.id) or Decimal("0")
            previous_balance_due = max(student.previous_balance - previous_balance_paid, Decimal("0"))
            total_due = due_monthly_total + due_transport_total + previous_balance_due

            if total_due > 0:
                results.append(
                    {
                        "student_id": student.id,
                        "student_name": student.name,
                        "registration_number": student.registration_number,
                        "father_name": student.father_name,
                        "grandfather_name": student.grandfather_name,
                        "phone": student.phone,
                        "class_id": student.school_class_id,
                        "class_name": student.school_class.name,
                        "class_year_shamsi": student.school_class.year_shamsi,
                        "expected_monthly_fee": str(expected_monthly),
                        "paid_monthly_fee": str(paid_monthly_total),
                        # Monthly fee dues (previous vs current vs total)
                        "due_monthly_fee_previous": str(due_monthly_previous),
                        "due_monthly_fee_current": str(due_monthly_current),
                        "due_monthly_previous_months_count": monthly_previous_unpaid_months,
                        "due_monthly_fee": str(due_monthly_total),
                        "expected_transport_fee": str(expected_transport),
                        "paid_transport_fee": str(paid_transport_total),
                        # Transport dues (previous vs current vs total)
                        "due_transport_fee_previous": str(due_transport_previous),
                        "due_transport_fee_current": str(due_transport_current),
                        "due_transport_previous_months_count": transport_previous_unpaid_months,
                        "due_transport_fee": str(due_transport_total),
                        "previous_balance": _money_str(student.previous_balance),
                        "paid_previous_balance": _money_str(previous_balance_paid),
                        "due_previous_balance": _money_str(previous_balance_due),
                        "due_amount": str(total_due),
                    }
                )

        return Response(
            {
                "month_shamsi": month_shamsi,
                "dues_from_month_shamsi": dues_from_month_shamsi,
                "months_count": len(months_in_range),
                "total_due_students": len(results),
                "results": results,
            }
        )


class ClassMonthlyFeesReportView(APIView):
    """
    Per-class totals for one Shamsi month: expected fees, paid, remaining, student counts.

    *free_students* counts students whose expected monthly and transport fees are both zero
    (fee-exempt for this snapshot).
    """

    def get(self, request):
        month_shamsi = request.query_params.get("month_shamsi")
        if not month_shamsi:
            year = request.query_params.get("year")
            month = request.query_params.get("month")
            if year and month:
                try:
                    month_shamsi = f"{int(year):04d}-{int(month):02d}"
                except ValueError:
                    return Response(
                        {"detail": "Invalid year/month format."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        if not month_shamsi:
            return Response(
                {"detail": "month_shamsi is required (YYYY-MM)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            _parse_shamsi_month(month_shamsi)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        monthly_fee_types = FeeType.objects.filter(name__icontains="monthly")
        if not monthly_fee_types.exists():
            return Response(
                {"detail": "Monthly FeeType not found. Create a FeeType with 'monthly' in its name."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        transport_fee_types = FeeType.objects.filter(name__icontains="transport")

        rows = []
        for cls in SchoolClass.objects.all().order_by("year_shamsi", "name"):
            students = list(Student.objects.filter(school_class=cls))
            n = len(students)
            if n == 0:
                rows.append(
                    {
                        "class_id": cls.id,
                        "class_name": cls.name,
                        "year_shamsi": cls.year_shamsi,
                        "class_label": f"{cls.name} ({cls.year_shamsi})",
                        "student_count": 0,
                        "total_monthly_expected": "0",
                        "total_transport_expected": "0",
                        "total_monthly_paid": "0",
                        "total_transport_paid": "0",
                        "total_monthly_remaining": "0",
                        "total_transport_remaining": "0",
                        "free_students_count": 0,
                    }
                )
                continue

            sids = [s.id for s in students]
            monthly_paid_map = {
                row["student_id"]: row["paid"]
                for row in Payment.objects.filter(
                    student_id__in=sids,
                    month_shamsi=month_shamsi,
                    fee_type__in=monthly_fee_types,
                )
                .values("student_id")
                .annotate(paid=Sum("amount"))
            }
            transport_paid_map = {
                row["student_id"]: row["paid"]
                for row in Payment.objects.filter(
                    student_id__in=sids,
                    month_shamsi=month_shamsi,
                    fee_type__in=transport_fee_types,
                )
                .values("student_id")
                .annotate(paid=Sum("amount"))
            }

            total_m_exp = Decimal("0")
            total_t_exp = Decimal("0")
            total_m_paid = Decimal("0")
            total_t_paid = Decimal("0")
            total_m_rem = Decimal("0")
            total_t_rem = Decimal("0")
            free_students = 0

            for s in students:
                exp_m = (
                    s.monthly_fee_override if s.monthly_fee_override is not None else cls.monthly_fee
                )
                exp_t = (
                    s.transport_fee_override if s.transport_fee_override is not None else cls.transport_fee
                )
                paid_m = monthly_paid_map.get(s.id) or Decimal("0")
                paid_t = transport_paid_map.get(s.id) or Decimal("0")
                total_m_exp += exp_m
                total_t_exp += exp_t
                total_m_paid += paid_m
                total_t_paid += paid_t
                total_m_rem += max(exp_m - paid_m, Decimal("0"))
                total_t_rem += max(exp_t - paid_t, Decimal("0"))
                if exp_m == 0 and exp_t == 0:
                    free_students += 1

            rows.append(
                {
                    "class_id": cls.id,
                    "class_name": cls.name,
                    "year_shamsi": cls.year_shamsi,
                    "class_label": f"{cls.name} ({cls.year_shamsi})",
                    "student_count": n,
                    "total_monthly_expected": str(total_m_exp),
                    "total_transport_expected": str(total_t_exp),
                    "total_monthly_paid": str(total_m_paid),
                    "total_transport_paid": str(total_t_paid),
                    "total_monthly_remaining": str(total_m_rem),
                    "total_transport_remaining": str(total_t_rem),
                    "free_students_count": free_students,
                }
            )

        return Response({"month_shamsi": month_shamsi, "classes": rows})


class TeacherStatementReportView(APIView):
    """
    Printable per-teacher salary statement.
    """

    def get(self, request):
        teacher_id = request.query_params.get("teacher_id")
        if not teacher_id:
            return Response({"detail": "teacher_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        requested_month_shamsi = request.query_params.get("month_shamsi") or _current_shamsi_month()
        try:
            month_shamsi = _cap_teacher_salary_month(requested_month_shamsi)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            teacher = Teacher.objects.get(pk=teacher_id)
        except Teacher.DoesNotExist:
            return Response({"detail": "Teacher not found."}, status=status.HTTP_404_NOT_FOUND)

        start_month_shamsi = _teacher_started_month_shamsi(teacher)
        start_year, start_month = _parse_shamsi_month(start_month_shamsi)
        end_year, end_month = _parse_shamsi_month(month_shamsi)
        if (start_year, start_month) > (end_year, end_month):
            months_in_scope = []
        else:
            try:
                months_in_scope = _iter_shamsi_months(start_month_shamsi, month_shamsi)
            except ValueError as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payment_qs = TeacherSalaryPayment.objects.filter(
            teacher_id=teacher.id,
            month_shamsi__in=months_in_scope,
        ).order_by("date_shamsi", "id")
        payments = list(payment_qs)

        paid_by_month = {
            row["month_shamsi"]: row["paid"]
            for row in payment_qs.values("month_shamsi").annotate(paid=Sum("amount")).order_by("month_shamsi")
        }

        month_rows = []
        total_should_pay = Decimal("0")
        total_paid = Decimal("0")
        total_balance = Decimal("0")
        for month in months_in_scope:
            expected = teacher.salary
            paid = paid_by_month.get(month) or Decimal("0")
            due = max(expected - paid, Decimal("0"))
            month_rows.append(
                {
                    "month_shamsi": month,
                    "expected_salary": _money_str(expected),
                    "paid_salary": _money_str(paid),
                    "due_salary": _money_str(due),
                }
            )
            total_should_pay += expected
            total_paid += paid
            total_balance += due

        return Response(
            {
                "teacher": {
                    "id": teacher.id,
                    "name": teacher.name,
                    "father_name": teacher.father_name,
                    "phone": teacher.phone,
                    "email": teacher.email,
                    "address": teacher.address,
                    "department": teacher.department,
                    "salary": _money_str(teacher.salary),
                    "created_date_shamsi": _teacher_created_date_shamsi(teacher),
                    "start_month_shamsi": start_month_shamsi,
                },
                "through_month_shamsi": month_shamsi,
                "requested_month_shamsi": requested_month_shamsi,
                "months_count": len(months_in_scope),
                "months": month_rows,
                "salary_payments": TeacherSalaryPaymentSerializer(payments, many=True).data,
                "summary": {
                    "total_expected": _money_str(total_should_pay),
                    "total_paid": _money_str(total_paid),
                    "total_balance": _money_str(total_balance),
                    "total_due": _money_str(total_balance),
                },
            }
        )


class ExpenseCategoryStatementReportView(APIView):
    """
    Printable per-category expense statement with optional date range.
    """

    def get(self, request):
        category_id = request.query_params.get("category_id")
        if not category_id:
            return Response({"detail": "category_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            category = ExpenseCategory.objects.get(pk=category_id)
        except ExpenseCategory.DoesNotExist:
            return Response({"detail": "Expense category not found."}, status=status.HTTP_404_NOT_FOUND)

        expense_qs = Expense.objects.select_related("category").filter(category_id=category.id).order_by("date_shamsi", "id")

        start = request.query_params.get("start")
        end = request.query_params.get("end")
        month_shamsi = request.query_params.get("month_shamsi")

        if start and end:
            try:
                start_date = _parse_shamsi_date(start)
                end_date = _parse_shamsi_date(end)
            except ValueError as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
            expense_qs = expense_qs.filter(date_shamsi__gte=start_date, date_shamsi__lte=end_date)
        elif month_shamsi:
            try:
                year, month = _parse_shamsi_month(month_shamsi)
            except ValueError as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
            expense_qs = expense_qs.filter(date_shamsi__year=year, date_shamsi__month=month)

        expenses = list(expense_qs)
        total_amount = expense_qs.aggregate(total=Sum("amount")).get("total") or Decimal("0")

        return Response(
            {
                "category": {
                    "id": category.id,
                    "name": category.name,
                },
                "filters": {
                    "start": start or "",
                    "end": end or "",
                    "month_shamsi": month_shamsi or "",
                },
                "summary": {
                    "total_amount": _money_str(total_amount),
                    "expenses_count": len(expenses),
                },
                "expenses": ExpenseSerializer(expenses, many=True).data,
            }
        )


class StudentStatementReportView(APIView):
    """
    Printable per-student statement with recurring fees, one-time items, and all payments.
    """

    def get(self, request):
        student_id = request.query_params.get("student_id")
        if not student_id:
            return Response(
                {"detail": "student_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        month_shamsi = request.query_params.get("month_shamsi") or _current_shamsi_month()
        try:
            _parse_shamsi_month(month_shamsi)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related("school_class").get(pk=student_id)
        except Student.DoesNotExist:
            return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        if student.school_class is None:
            return Response(
                {"detail": "Student is not assigned to a class."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        enrolled_month_shamsi = _student_enrolled_month_shamsi(student)
        start_month_shamsi = max(
            enrolled_month_shamsi,
            f"{student.school_class.year_shamsi}-01",
        )
        try:
            months_in_scope = _iter_shamsi_months(start_month_shamsi, month_shamsi)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        monthly_fee_types = FeeType.objects.filter(name__icontains="monthly")
        transport_fee_types = FeeType.objects.filter(name__icontains="transport")
        uniform_fee_types = FeeType.objects.filter(name__icontains="uniform")
        book_fee_types = FeeType.objects.filter(name__icontains="book")
        previous_balance_fee_types = _previous_balance_fee_types()

        payment_qs = Payment.objects.select_related("student", "fee_type", "school_class").filter(
            student_id=student.id,
            month_shamsi__in=months_in_scope,
        )
        payments = list(payment_qs.order_by("date_shamsi", "id"))
        payment_serializer = PaymentSerializer(payments, many=True)

        monthly_paid_map = _payment_totals_by_month(payment_qs, monthly_fee_types)
        transport_paid_map = _payment_totals_by_month(payment_qs, transport_fee_types)

        recurring_monthly_fee = _student_fee_value(student.monthly_fee_override, student.school_class.monthly_fee)
        recurring_transport_fee = _student_fee_value(student.transport_fee_override, student.school_class.transport_fee)
        one_time_uniform_fee = _student_fee_value(student.uniform_fee_override, student.school_class.uniform_fee)
        one_time_book_fee = _student_fee_value(student.book_fee_override, student.school_class.book_fee)

        month_rows = []
        total_monthly_expected = Decimal("0")
        total_monthly_paid = Decimal("0")
        total_monthly_due = Decimal("0")
        total_transport_expected = Decimal("0")
        total_transport_paid = Decimal("0")
        total_transport_due = Decimal("0")

        for month in months_in_scope:
            paid_monthly = monthly_paid_map.get(month) or Decimal("0")
            paid_transport = transport_paid_map.get(month) or Decimal("0")
            due_monthly = max(recurring_monthly_fee - paid_monthly, Decimal("0"))
            due_transport = max(recurring_transport_fee - paid_transport, Decimal("0"))
            month_rows.append(
                {
                    "month_shamsi": month,
                    "expected_monthly_fee": _money_str(recurring_monthly_fee),
                    "paid_monthly_fee": _money_str(paid_monthly),
                    "due_monthly_fee": _money_str(due_monthly),
                    "expected_transport_fee": _money_str(recurring_transport_fee),
                    "paid_transport_fee": _money_str(paid_transport),
                    "due_transport_fee": _money_str(due_transport),
                    "total_due": _money_str(due_monthly + due_transport),
                }
            )
            total_monthly_expected += recurring_monthly_fee
            total_monthly_paid += paid_monthly
            total_monthly_due += due_monthly
            total_transport_expected += recurring_transport_fee
            total_transport_paid += paid_transport
            total_transport_due += due_transport

        fee_totals = [
            {
                "fee_type_name": row["fee_type__name"],
                "total_paid": _money_str(row["paid"]),
            }
            for row in (
                payment_qs.values("fee_type__name")
                .annotate(paid=Sum("amount"))
                .order_by("fee_type__name")
            )
        ]

        uniform_paid_total = _payment_total_for_types(payment_qs, uniform_fee_types)
        book_paid_total = _payment_total_for_types(payment_qs, book_fee_types)
        previous_balance_paid_total = _payment_total_for_types(payment_qs, previous_balance_fee_types)
        other_paid_total = sum(
            Decimal(str(row["paid"] or Decimal("0")))
            for row in fee_totals
            if not _is_statement_fee_type_name(row["fee_type_name"])
        )

        uniform_due = max(one_time_uniform_fee - uniform_paid_total, Decimal("0"))
        book_due = max(one_time_book_fee - book_paid_total, Decimal("0"))
        previous_balance_due = max(student.previous_balance - previous_balance_paid_total, Decimal("0"))
        recurring_due = total_monthly_due + total_transport_due
        one_time_due = uniform_due + book_due + previous_balance_due
        total_paid = (
            total_monthly_paid
            + total_transport_paid
            + uniform_paid_total
            + book_paid_total
            + previous_balance_paid_total
            + other_paid_total
        )
        total_expected = (
            total_monthly_expected
            + total_transport_expected
            + one_time_uniform_fee
            + one_time_book_fee
            + student.previous_balance
        )
        total_balance = total_expected - total_paid

        return Response(
            {
                "student": {
                    "id": student.id,
                    "name": student.name,
                    "registration_number": student.registration_number,
                    "father_name": student.father_name,
                    "grandfather_name": student.grandfather_name,
                    "phone": student.phone,
                    "class_id": student.school_class_id,
                    "class_name": student.school_class.name,
                    "class_year_shamsi": student.school_class.year_shamsi,
                    "enrolled_date_shamsi": _student_enrolled_date_shamsi(student),
                    "enrolled_month_shamsi": enrolled_month_shamsi,
                    "previous_balance": _money_str(student.previous_balance),
                },
                "through_month_shamsi": month_shamsi,
                "start_month_shamsi": start_month_shamsi,
                "months_count": len(months_in_scope),
                "months": month_rows,
                "payments": payment_serializer.data,
                "fee_totals": fee_totals,
                "summary": {
                    "total_expected": _money_str(total_expected),
                    "monthly_expected": _money_str(total_monthly_expected),
                    "monthly_paid": _money_str(total_monthly_paid),
                    "monthly_due": _money_str(total_monthly_due),
                    "transport_expected": _money_str(total_transport_expected),
                    "transport_paid": _money_str(total_transport_paid),
                    "transport_due": _money_str(total_transport_due),
                    "uniform_expected": _money_str(one_time_uniform_fee),
                    "uniform_paid": _money_str(uniform_paid_total),
                    "uniform_due": _money_str(uniform_due),
                    "book_expected": _money_str(one_time_book_fee),
                    "book_paid": _money_str(book_paid_total),
                    "book_due": _money_str(book_due),
                    "previous_balance_expected": _money_str(student.previous_balance),
                    "previous_balance_paid": _money_str(previous_balance_paid_total),
                    "previous_balance_due": _money_str(previous_balance_due),
                    "other_paid": _money_str(other_paid_total),
                    "recurring_due": _money_str(recurring_due),
                    "one_time_due": _money_str(one_time_due),
                    "total_paid": _money_str(total_paid),
                    "total_due": _money_str(total_balance),
                    "total_balance": _money_str(total_balance),
                },
            }
        )


_BACKUP_MAX_UPLOAD_BYTES = 50 * 1024 * 1024


class BackupExportView(APIView):
    """
    Download full school data as JSON (users, API tokens, core app).
    Same format as `python manage.py export_backup <file>`.
    """

    def get(self, request):
        raw = build_dumpdata_backup_json().encode("utf-8")
        filename = f"school_rasool_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        response = HttpResponse(raw, content_type="application/json; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class BackupRestoreView(APIView):
    """
    Upload a backup JSON file. Clears the database, then loads the fixture.
    Requires multipart field `confirm` = RESTORE and `file` = backup .json
    """

    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if request.data.get("confirm") != "RESTORE":
            return Response(
                {
                    "detail": "Restore refused. Send form field confirm=RESTORE (exact text) together with the file.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Missing file field."}, status=status.HTTP_400_BAD_REQUEST)

        body = upload.read()
        if len(body) > _BACKUP_MAX_UPLOAD_BYTES:
            return Response({"detail": "Backup file is too large."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            decoded = body.decode("utf-8")
            parsed = json.loads(decoded)
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            return Response(
                {"detail": f"File is not valid UTF-8 JSON: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not isinstance(parsed, list) or (parsed and not isinstance(parsed[0], dict)):
            return Response(
                {"detail": "Invalid backup format (expected a JSON array)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if parsed and "model" not in parsed[0]:
            return Response(
                {"detail": "Invalid backup format (missing model keys)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        repair_backup_fixture_shamsi_dates(parsed)
        fixed_body = json.dumps(parsed, indent=2, ensure_ascii=False).encode("utf-8")

        fd, path = tempfile.mkstemp(suffix=".json")
        try:
            with os.fdopen(fd, "wb") as tmp:
                tmp.write(fixed_body)
            try:
                call_command("flush", interactive=False)
            except Exception as exc:  # noqa: BLE001
                return Response(
                    {"detail": f"Could not clear database: {exc}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            try:
                call_command("loaddata", path)
            except Exception as exc:  # noqa: BLE001
                return Response(
                    {
                        "detail": (
                            "Restore failed after the database was cleared. "
                            f"Re-import this or another backup file. Error: {exc}"
                        ),
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass

        return Response({"detail": "Backup restored successfully. Log in again if your session was reset."})


def _parse_shamsi_date(value: str) -> jdatetime.date:
    try:
        year, month, day = [int(part) for part in value.split("-")]
    except ValueError as exc:
        raise ValueError("Invalid date format. Use YYYY-MM-DD.") from exc
    return jdatetime.date(year, month, day)


def _parse_shamsi_month(value: str) -> tuple[int, int]:
    try:
        year, month = [int(part) for part in value.split("-")]
    except ValueError as exc:
        raise ValueError("Invalid month format. Use YYYY-MM.") from exc
    return year, month


def _parse_shamsi_year(value: str) -> int:
    try:
        return int(value)
    except ValueError as exc:
        raise ValueError("Invalid year format. Use YYYY.") from exc


def _current_shamsi_month() -> str:
    today = jdatetime.date.today()
    return f"{today.year:04d}-{today.month:02d}"


def _teacher_started_month_shamsi(teacher: Teacher) -> str:
    created = jdatetime.datetime.fromgregorian(datetime=timezone.localtime(teacher.created_at))
    return f"{created.year:04d}-{created.month:02d}"


def _teacher_created_date_shamsi(teacher: Teacher) -> str:
    created = jdatetime.datetime.fromgregorian(datetime=timezone.localtime(teacher.created_at))
    return f"{created.year:04d}-{created.month:02d}-{created.day:02d}"


def _cap_teacher_salary_month(month_shamsi: str) -> str:
    year, month = _parse_shamsi_month(month_shamsi)
    if month > 9:
        month = 9
    return f"{year:04d}-{month:02d}"


def _student_enrolled_date_shamsi(student: Student) -> str:
    enrolled = jdatetime.datetime.fromgregorian(datetime=timezone.localtime(student.created_at))
    return f"{enrolled.year:04d}-{enrolled.month:02d}-{enrolled.day:02d}"


def _student_enrolled_month_shamsi(student: Student) -> str:
    enrolled = jdatetime.datetime.fromgregorian(datetime=timezone.localtime(student.created_at))
    return f"{enrolled.year:04d}-{enrolled.month:02d}"


def _iter_shamsi_months(start_month: str, end_month: str) -> list[str]:
    start_year, start_num = _parse_shamsi_month(start_month)
    end_year, end_num = _parse_shamsi_month(end_month)
    if (start_year, start_num) > (end_year, end_num):
        raise ValueError("Start month cannot be after the end month.")

    months = []
    year, month = start_year, start_num
    while (year, month) <= (end_year, end_num):
        months.append(f"{year:04d}-{month:02d}")
        month += 1
        if month > 12:
            year += 1
            month = 1
    return months


def _student_fee_value(override, class_value):
    return override if override is not None else class_value


def _money_str(value) -> str:
    amount = Decimal("0") if value in (None, "") else Decimal(str(value))
    return str(amount.quantize(Decimal("0.00")))


def _payment_totals_by_month(payment_qs, fee_types):
    rows = {}
    if not fee_types.exists():
        return rows
    for row in (
        payment_qs.filter(fee_type__in=fee_types)
        .values("month_shamsi")
        .annotate(paid=Sum("amount"))
        .order_by("month_shamsi")
    ):
        rows[row["month_shamsi"]] = row["paid"] or Decimal("0")
    return rows


def _payment_total_for_types(payment_qs, fee_types):
    if not fee_types.exists():
        return Decimal("0")
    total = payment_qs.filter(fee_type__in=fee_types).aggregate(paid=Sum("amount")).get("paid")
    return total or Decimal("0")


def _is_statement_fee_type_name(name: str) -> bool:
    lowered = (name or "").lower()
    return any(key in lowered for key in ("monthly", "transport", "uniform", "book", "previous balance"))


def _previous_balance_fee_types():
    return FeeType.objects.filter(name__iexact="Previous Balance")


def _normalize_headers(headers):
    return [str(h or "").strip().lower() for h in headers]


def _read_student_rows(upload):
    name = (upload.name or "").lower()
    if name.endswith(".csv"):
        text = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV file is missing header row.")
        normalized = _normalize_headers(reader.fieldnames)
        rows = []
        for idx, row in enumerate(reader, start=2):
            mapped = {normalized[i]: (v.strip() if isinstance(v, str) else v) for i, v in enumerate(row.values())}
            rows.append((idx, mapped))
        return rows

    if name.endswith(".xlsx"):
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration as exc:
            raise ValueError("Excel file is empty.") from exc
        headers = _normalize_headers(raw_headers)
        if not any(headers):
            raise ValueError("Excel file is missing header row.")
        rows = []
        for idx, values in enumerate(iterator, start=2):
            mapped = {}
            for i, key in enumerate(headers):
                if not key:
                    continue
                value = values[i] if i < len(values) else ""
                mapped[key] = value.strip() if isinstance(value, str) else value
            rows.append((idx, mapped))
        return rows

    raise ValueError("Unsupported file format. Use .csv or .xlsx.")


def _parse_optional_decimal(value, field_name, row_number, errors):
    if value in (None, ""):
        return None
    try:
        parsed = Decimal(str(value))
    except Exception:  # noqa: BLE001
        errors.append({"row": row_number, "field": field_name, "message": "Must be a number."})
        return None
    if parsed < 0:
        errors.append({"row": row_number, "field": field_name, "message": "Must be >= 0."})
        return None
    return parsed


def _build_student_from_row(row, row_number, classes_by_id, classes_by_name_year):
    errors = []
    required_fields = ["name", "registration_number", "father_name", "grandfather_name", "phone"]
    for field in required_fields:
        if not str(row.get(field, "")).strip():
            errors.append({"row": row_number, "field": field, "message": "This field is required."})

    phone = str(row.get("phone", "")).strip()
    if phone and not phone.isdigit():
        errors.append({"row": row_number, "field": "phone", "message": "Only digits are allowed."})

    school_class = None
    class_id = str(row.get("class_id", "") or "").strip()
    class_name = str(row.get("class_name", "") or "").strip()
    year_shamsi = str(row.get("year_shamsi", "") or "").strip()

    if class_id:
        school_class = classes_by_id.get(class_id)
        if not school_class:
            errors.append({"row": row_number, "field": "class_id", "message": "Class not found."})
    elif class_name and year_shamsi:
        school_class = classes_by_name_year.get((class_name.lower(), year_shamsi))
        if not school_class:
            errors.append(
                {
                    "row": row_number,
                    "field": "class_name",
                    "message": "Class not found for given class_name + year_shamsi.",
                }
            )
    elif class_name or year_shamsi:
        errors.append(
            {
                "row": row_number,
                "field": "class_name/year_shamsi",
                "message": "Provide both class_name and year_shamsi together.",
            }
        )

    monthly_fee_override = _parse_optional_decimal(
        row.get("monthly_fee_override"), "monthly_fee_override", row_number, errors
    )
    transport_fee_override = _parse_optional_decimal(
        row.get("transport_fee_override"), "transport_fee_override", row_number, errors
    )
    uniform_fee_override = _parse_optional_decimal(
        row.get("uniform_fee_override"), "uniform_fee_override", row_number, errors
    )
    book_fee_override = _parse_optional_decimal(
        row.get("book_fee_override"), "book_fee_override", row_number, errors
    )
    previous_balance = _parse_optional_decimal(
        row.get("previous_balance"), "previous_balance", row_number, errors
    )

    if errors:
        return None, errors

    student = Student(
        school_class=school_class,
        name=str(row.get("name", "")).strip(),
        registration_number=str(row.get("registration_number", "")).strip(),
        father_name=str(row.get("father_name", "")).strip(),
        grandfather_name=str(row.get("grandfather_name", "")).strip(),
        phone=phone,
        monthly_fee_override=monthly_fee_override,
        transport_fee_override=transport_fee_override,
        uniform_fee_override=uniform_fee_override,
        book_fee_override=book_fee_override,
        previous_balance=previous_balance or Decimal("0"),
    )
    return student, []
